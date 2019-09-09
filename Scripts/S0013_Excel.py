import openpyxl
import os

os.chdir(os.path.join(os.getcwd(), "Sources"))
wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))

print('\n---------------Getting Sheets---------------------')
print('wb.sheetnames', wb.sheetnames)
sheet3 = wb.worksheets[0]
print('worksheets[0].title', sheet3.title)
print('wb.active.title', wb.active.title)
print('sheet3.max_column', sheet3.max_column)
print(r"sheet3['A1':'B3']", sheet3['A1':'B3'])
print("sheet3['B$']", sheet3['B$'])
print("list(sheet3.columns)[1]", list(sheet3.columns)[2])

print('\n---------------Getting Cells---------------------')
cell1 = wb.active.cell(2, 5)
print(type(cell1.value))
print('cell1.value', cell1.value)
print(cell1.row, cell1.column)

print('\n--------------utils-get_column_letter-column_index_from_string--------------------')
from openpyxl.utils import get_column_letter, column_index_from_string
print('get_column_letter(1)', get_column_letter(28))
print(r"column_index_from_string('AA')", column_index_from_string('AA'))


print('\n---------------Creating and Saving---------------------')
wb = openpyxl.Workbook()
wb.worksheets[0].title = 'Mysheet1'
wb.create_sheet("My_sheet2")
wb.create_sheet("My_sheet3", 0)
wb.create_sheet("My_sheet4")
wb.save("workbook_hank.xlsx")
wb = openpyxl.load_workbook("workbook_hank.xlsx")
wb.remove(wb.worksheets[3])
for x in range(0,50):
    wb.worksheets[0].cell(row=x+1,column=1).value = "values_" + str(x)
wb.save("workbook_hank.xlsx")

print('\n------------------Setting the Font Style of Cells------------------')
from openpyxl.styles import Font
wb = openpyxl.load_workbook("workbook_hank.xlsx")
wb.worksheets[0].insert_rows(0,1)
wb.worksheets[0].cell(row=1,column=1).value = "Col1"
wb.worksheets[0].cell(row=1,column=1).font = Font(size=32,bold=True,italic=True)
wb.save("workbook_hank.xlsx")

print('\n---------------Adjusting Rows and Columns---------------------')
wb = openpyxl.load_workbook("workbook_hank.xlsx")
wb.worksheets[0].row_dimensions[2].height = 80
wb.worksheets[0].column_dimensions["B"].width = 100
wb.save("workbook_hank.xlsx")

print('\n---------------Merging and Unmerging Cells---------------------')
wb = openpyxl.load_workbook("workbook_hank.xlsx")
wb.worksheets[0].merge_cells("A1:B1")
wb.worksheets[0].merge_cells("A2:C3")
wb.worksheets[0].unmerge_cells("A1:B1")
wb.save("workbook_hank.xlsx")